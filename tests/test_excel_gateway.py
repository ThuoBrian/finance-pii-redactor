"""Tests for the Excel write adapter: highlighting plus the Crosswalk sheet.

Framework-free besides pandas/openpyxl themselves: no Presidio/spaCy involved,
so these run without the language model (matching the rest of the suite).
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from finance_redactor.infrastructure.documents.excel_gateway import (
    OpenpyxlExcelGateway,
)

_CROSSWALK_COLUMNS = [
    "Original name",
    "Entity type",
    "Category",
    "Pseudonym",
    "Flagged",
    "Possible match",
]


def _crosswalk_df(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=_CROSSWALK_COLUMNS)


def test_write_produces_redacted_and_crosswalk_sheets():
    df = pd.DataFrame({"notes": ["Paid to STF-91345"]})
    crosswalk_df = _crosswalk_df(
        [
            {
                "Original name": "Jane Doe",
                "Entity type": "PERSON",
                "Category": "Staff",
                "Pseudonym": "STF-91345",
                "Flagged": "",
                "Possible match": "",
            }
        ]
    )

    result = OpenpyxlExcelGateway().write(df, set(), crosswalk_df)

    workbook = load_workbook(BytesIO(result))
    assert workbook.sheetnames == ["Redacted", "Crosswalk"]


def test_redacted_sheet_content_and_highlighting_are_preserved():
    df = pd.DataFrame({"notes": ["Paid to STF-91345", "no name here"]})

    result = OpenpyxlExcelGateway().write(df, {(0, "notes")}, _crosswalk_df([]))

    workbook = load_workbook(BytesIO(result))
    sheet = workbook["Redacted"]
    # Header row, then two data rows.
    assert sheet.cell(row=1, column=1).value == "notes"
    assert sheet.cell(row=2, column=1).value == "Paid to STF-91345"
    assert sheet.cell(row=3, column=1).value == "no name here"
    # Row 0 ("notes" -> column 1) was flagged as changed; row 1 was not.
    assert sheet.cell(row=2, column=1).fill.fgColor.rgb == "00FFFF00"
    assert sheet.cell(row=3, column=1).fill.fgColor.rgb in (None, "00000000")


def test_crosswalk_sheet_matches_the_passed_dataframe():
    df = pd.DataFrame({"notes": ["x"]})
    crosswalk_df = _crosswalk_df(
        [
            {
                "Original name": "Jane Doe",
                "Entity type": "PERSON",
                "Category": "Staff",
                "Pseudonym": "STF-91345",
                "Flagged": "",
                "Possible match": "",
            },
            {
                "Original name": "Micheal Mugo",
                "Entity type": "PERSON",
                "Category": "",
                "Pseudonym": "PSN-AUTO-0BA3D",
                "Flagged": "yes",
                "Possible match": "Michael Mugo (STF-12345, 92% match)",
            },
        ]
    )

    result = OpenpyxlExcelGateway().write(df, set(), crosswalk_df)

    read_back = pd.read_excel(BytesIO(result), sheet_name="Crosswalk")
    assert list(read_back.columns) == _CROSSWALK_COLUMNS
    assert read_back.iloc[0]["Original name"] == "Jane Doe"
    assert read_back.iloc[1]["Flagged"] == "yes"
    # Header row is bolded for readability.
    sheet = load_workbook(BytesIO(result))["Crosswalk"]
    assert sheet.cell(row=1, column=1).font.bold is True


def test_write_handles_an_empty_crosswalk_without_crashing():
    df = pd.DataFrame({"notes": ["no PII here"]})

    result = OpenpyxlExcelGateway().write(df, set(), _crosswalk_df([]))

    read_back = pd.read_excel(BytesIO(result), sheet_name="Crosswalk")
    assert list(read_back.columns) == _CROSSWALK_COLUMNS
    assert len(read_back) == 0


def _round_tripped_workbook_bytes(data: dict) -> bytes:
    """Write ``data`` to real xlsx bytes and back, like an actual upload.

    Constructing a DataFrame directly in memory (``pd.DataFrame({...})``)
    does not reproduce the dtype a real upload gets: pandas 3.0 infers its
    own string dtype when *reading* an xlsx file, which can differ from
    whatever dtype an in-memory literal happens to get. Round-tripping
    through real bytes via ``gateway.read()`` is what actually exercises
    ``text_columns()`` the way the app does.
    """
    buffer = BytesIO()
    pd.DataFrame(data).to_excel(buffer, index=False)
    buffer.seek(0)
    return buffer.getvalue()


def test_text_columns_includes_string_columns_from_a_real_upload():
    # Regression test: pandas 3.0 infers a dedicated string dtype (shown as
    # "str" in df.dtypes) for text columns read from an xlsx file, not the
    # legacy "object" dtype - a `dtype == object` check silently matched
    # nothing, so no column was ever pre-selected for a real uploaded file.
    gateway = OpenpyxlExcelGateway()
    df = gateway.read(
        BytesIO(
            _round_tripped_workbook_bytes(
                {"Description": ["Consulting payment"], "Payee": ["Jane Doe"]}
            )
        )
    )

    assert set(gateway.text_columns(df)) == {"Description", "Payee"}


def test_text_columns_excludes_numeric_and_date_columns_from_a_real_upload():
    gateway = OpenpyxlExcelGateway()
    df = gateway.read(
        BytesIO(
            _round_tripped_workbook_bytes(
                {
                    "Payee": ["Jane Doe"],
                    "Amount": [1000],
                    "Date": pd.to_datetime(["2020-01-01"]),
                }
            )
        )
    )

    assert gateway.text_columns(df) == ["Payee"]
