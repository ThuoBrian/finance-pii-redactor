"""Excel read/write adapter (pandas + openpyxl).

Implements the :class:`ExcelGateway` port. Encapsulates all spreadsheet I/O,
including the yellow highlighting of changed cells, which previously lived in
the UI module.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

_HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
_HEADER_FONT = Font(bold=True)
_SHEET_NAME = "Redacted"
_CROSSWALK_SHEET_NAME = "Crosswalk"


class OpenpyxlExcelGateway:
    """Reads workbooks into DataFrames and writes highlighted redacted copies."""

    def read(self, source: object) -> pd.DataFrame:
        """Load the first sheet of a workbook into a DataFrame."""
        return pd.read_excel(source, engine="openpyxl")

    def text_columns(self, df: pd.DataFrame) -> list[str]:
        """Return free-text (string dtype) columns, used as the scan default.

        Uses ``pandas.api.types.is_string_dtype`` rather than checking
        ``dtype == object``: pandas 3.0 infers a dedicated string dtype
        (``StringDtype``, shown as ``str`` in ``df.dtypes``) for text columns
        read from Excel/CSV by default, not the legacy ``object`` dtype an
        `== object` check was written against - that check silently matched
        zero columns on pandas 3.0, so no column was ever pre-selected here.
        ``is_string_dtype`` matches both the new string dtype and legacy
        object-dtype string columns, while still excluding numeric/date
        columns and an all-null object column.
        """
        return [col for col in df.columns if pd.api.types.is_string_dtype(df[col])]

    def write(
        self,
        df: pd.DataFrame,
        highlighted_cells: set[tuple[int, str]],
        crosswalk_df: pd.DataFrame,
    ) -> bytes:
        """Serialize ``df`` to xlsx bytes, highlighting the given cells yellow.

        ``crosswalk_df`` is written as a second "Crosswalk" sheet alongside
        the redacted data, so the name-to-pseudonym mapping travels with the
        file it describes.
        """
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=_SHEET_NAME)
            crosswalk_df.to_excel(writer, index=False, sheet_name=_CROSSWALK_SHEET_NAME)
        buffer.seek(0)

        workbook = load_workbook(buffer)
        worksheet = workbook[_SHEET_NAME]
        col_index = {col: idx + 1 for idx, col in enumerate(df.columns)}
        for row_idx, col_name in highlighted_cells:
            if col_name in col_index:
                # +2 accounts for the header row and 1-based indexing.
                worksheet.cell(
                    row=row_idx + 2, column=col_index[col_name]
                ).fill = _HIGHLIGHT_FILL

        crosswalk_sheet = workbook[_CROSSWALK_SHEET_NAME]
        for cell in crosswalk_sheet[1]:
            cell.font = _HEADER_FONT

        out = BytesIO()
        workbook.save(out)
        return out.getvalue()
